# attendance_bot/services/report_service.py
import os
from datetime import date, datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import select, func, and_
from sqlalchemy.ext.asyncio import AsyncSession

from models import User, AttendanceLog, Department, AttendanceSummary, LeaveRequest


class ReportService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.reports_dir = "/app/reports"
        os.makedirs(self.reports_dir, exist_ok=True)

    async def generate_excel_report(
        self,
        report_type: str,
        year: int,
        month: Optional[int] = None,
        department_id: Optional[int] = None,
        user_id: Optional[int] = None,
    ) -> str:
        if report_type == "monthly":
            return await self._generate_monthly_excel(year, month, department_id, user_id)
        elif report_type == "yearly":
            return await self._generate_yearly_excel(year, department_id, user_id)
        elif report_type == "department":
            return await self._generate_department_excel(year, month, department_id)
        else:
            raise ValueError(f"Unknown report type: {report_type}")

    async def _generate_monthly_excel(
        self, year: int, month: int, department_id: Optional[int], user_id: Optional[int]
    ) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = f"{year}-{month:02d} 考勤报表"

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws.merge_cells("A1:K1")
        ws["A1"] = f"{year}年{month}月 考勤报表"
        ws["A1"].font = Font(bold=True, size=16)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 30

        # Headers
        headers = ["工号", "姓名", "部门", "应出勤", "实出勤", "迟到", "早退", "缺勤", "外勤", "加班(小时)", "请假(天)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Query data
        query = select(
            User.employee_number,
            User.full_name,
            Department.name.label("department_name"),
            AttendanceSummary.total_working_days,
            AttendanceSummary.present_days,
            AttendanceSummary.late_days,
            AttendanceSummary.early_leave_days,
            AttendanceSummary.absent_days,
            AttendanceSummary.field_work_days,
            AttendanceSummary.overtime_hours,
            AttendanceSummary.leave_days,
        ).join(AttendanceSummary, AttendanceSummary.user_id == User.id)\
         .outerjoin(Department, Department.id == User.department_id)\
         .where(
            AttendanceSummary.year == year,
            AttendanceSummary.month == month
        )

        if department_id:
            query = query.where(User.department_id == department_id)
        if user_id:
            query = query.where(User.id == user_id)

        result = await self.db.execute(query)
        rows = result.all()

        # Data rows
        for row_idx, row in enumerate(rows, 4):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column].width = adjusted_width

        filename = f"attendance_monthly_{year}_{month:02d}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        filepath = os.path.join(self.reports_dir, filename)
        wb.save(filepath)
        return filepath

    async def _generate_yearly_excel(
        self, year: int, department_id: Optional[int], user_id: Optional[int]
    ) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = f"{year}年 年度考勤报表"

        # Similar structure to monthly but aggregated by year
        ws.merge_cells("A1:L1")
        ws["A1"] = f"{year}年 年度考勤报表"
        ws["A1"].font = Font(bold=True, size=16)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["工号", "姓名", "部门", "应出勤", "实出勤", "迟到", "早退", "缺勤", "外勤", "加班(小时)", "请假(天)", "出勤率%"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Aggregate yearly data
        query = select(
            User.employee_number,
            User.full_name,
            Department.name.label("department_name"),
            func.sum(AttendanceSummary.total_working_days).label("total_working"),
            func.sum(AttendanceSummary.present_days).label("total_present"),
            func.sum(AttendanceSummary.late_days).label("total_late"),
            func.sum(AttendanceSummary.early_leave_days).label("total_early"),
            func.sum(AttendanceSummary.absent_days).label("total_absent"),
            func.sum(AttendanceSummary.field_work_days).label("total_field"),
            func.sum(AttendanceSummary.overtime_hours).label("total_overtime"),
            func.sum(AttendanceSummary.leave_days).label("total_leave"),
        ).join(AttendanceSummary, AttendanceSummary.user_id == User.id)\
         .outerjoin(Department, Department.id == User.department_id)\
         .where(AttendanceSummary.year == year)\
         .group_by(User.id, Department.name)

        if department_id:
            query = query.where(User.department_id == department_id)
        if user_id:
            query = query.where(User.id == user_id)

        result = await self.db.execute(query)
        rows = result.all()

        for row_idx, row in enumerate(rows, 4):
            values = list(row)
            # Calculate attendance rate
            total_working = values[3] or 0
            total_present = values[4] or 0
            rate = round(total_present / total_working * 100, 2) if total_working > 0 else 0
            values.append(rate)

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.alignment = Alignment(horizontal="center")

        filename = f"attendance_yearly_{year}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        filepath = os.path.join(self.reports_dir, filename)
        wb.save(filepath)
        return filepath

    async def _generate_department_excel(
        self, year: int, month: Optional[int], department_id: Optional[int]
    ) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "部门考勤分析"

        ws.merge_cells("A1:H1")
        ws["A1"] = f"部门考勤分析报表 ({year}年)"
        ws["A1"].font = Font(bold=True, size=16)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["部门", "总人数", "应出勤", "实出勤", "迟到率%", "缺勤率%", "出勤率%", "加班(小时)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        query = select(
            Department.name,
            func.count(User.id).label("total_users"),
            func.sum(AttendanceSummary.total_working_days).label("total_working"),
            func.sum(AttendanceSummary.present_days).label("total_present"),
            func.sum(AttendanceSummary.late_days).label("total_late"),
            func.sum(AttendanceSummary.absent_days).label("total_absent"),
            func.sum(AttendanceSummary.overtime_hours).label("total_overtime"),
        ).join(User, User.department_id == Department.id)\
         .join(AttendanceSummary, AttendanceSummary.user_id == User.id)\
         .where(AttendanceSummary.year == year)\
         .group_by(Department.id)

        if month:
            query = query.where(AttendanceSummary.month == month)
        if department_id:
            query = query.where(Department.id == department_id)

        result = await self.db.execute(query)
        rows = result.all()

        for row_idx, row in enumerate(rows, 4):
            dept_name, total_users, total_working, total_present, total_late, total_absent, total_overtime = row

            late_rate = round(total_late / total_working * 100, 2) if total_working else 0
            absent_rate = round(total_absent / total_working * 100, 2) if total_working else 0
            attendance_rate = round(total_present / total_working * 100, 2) if total_working else 0

            values = [dept_name, total_users, total_working, total_present, late_rate, absent_rate, attendance_rate, total_overtime]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.alignment = Alignment(horizontal="center")

        filename = f"attendance_department_{year}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        filepath = os.path.join(self.reports_dir, filename)
        wb.save(filepath)
        return filepath

    async def generate_pdf_report(
        self,
        report_type: str,
        year: int,
        month: Optional[int] = None,
        department_id: Optional[int] = None,
        user_id: Optional[int] = None,
    ) -> str:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        filename = f"attendance_{report_type}_{year}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
        filepath = os.path.join(self.reports_dir, filename)

        doc = SimpleDocTemplate(filepath, pagesize=landscape(A4))
        elements = []

        styles = getSampleStyleSheet()
        title = Paragraph(f"<b>考勤报表 - {year}年</b>", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 20))

        # For simplicity, create a basic table
        data = [["工号", "姓名", "部门", "应出勤", "实出勤", "迟到", "早退", "缺勤"]]

        query = select(
            User.employee_number,
            User.full_name,
            Department.name,
            AttendanceSummary.total_working_days,
            AttendanceSummary.present_days,
            AttendanceSummary.late_days,
            AttendanceSummary.early_leave_days,
            AttendanceSummary.absent_days,
        ).join(AttendanceSummary).outerjoin(Department)\
         .where(AttendanceSummary.year == year)

        if month:
            query = query.where(AttendanceSummary.month == month)

        result = await self.db.execute(query)
        for row in result.all():
            data.append(list(row))

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))

        elements.append(table)
        doc.build(elements)

        return filepath
